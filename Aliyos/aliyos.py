import os
import sys
#########################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

toaddrs  = 'EfraimMKrug@gmail.com'

last = 0
notAliyah = 0
publicLines = 0
closedLines = 0
paymentLines = 0
aliyahCount = 0
totalLines = 0

accountArray = []
accountArray.append(dict())
server = 0

def openTrx():
    wb = load_workbook(basedir + '\\shulCloud\\transactions.xlsx')
    return wb

def openPPL():
    wb = load_workbook(basedir + '\\shulCloud\\people.xlsx')
    return wb

def getTrx(sheet):
    global last
    global notAliyah
    global publicLines
    global closedLines
    global paymentLines
    global aliyahCount
    global totalLines

    for r in range(2, sheet.max_row + 1):
        #must be logged as 'Aliyah'
        totalLines += 1
        if sheet.cell(row=r, column=6).value != "Aliyahs":
            notAliyah += 1
            continue
        #skip public accounts...
        if sheet.cell(row=r, column=14).value == 0:
            publicLines += 1
            continue
        #skip if the pledge was already paid
        if sheet.cell(row=r, column=11).value == "Closed":
            closedLines += 1
            continue
        #skip if the line item is payment
        if sheet.cell(row=r, column=2).value[0] == "P":
            paymentLines += 1
            continue

        if sheet.cell(row=r, column=14).value in accountArray[last]:
            accountArray.append(dict())
            last += 1

        aliyahCount += 1
        accountArray[last][sheet.cell(row=r, column=14).value] = []
        accountArray[last][sheet.cell(row=r, column=14).value].append(sheet.cell(row=r, column=3).value)  #name
        accountArray[last][sheet.cell(row=r, column=14).value].append(sheet.cell(row=r, column=8).value)  #notes
        accountArray[last][sheet.cell(row=r, column=14).value].append(sheet.cell(row=r, column=9).value)  #charge

def getEmail(sheet):
    for accounts in accountArray:
        for id in accounts:
            accounts[id].append([])
            for r in range(2, sheet.max_row + 1):
                if (sheet.cell(row=r, column=83).value == id or sheet.cell(row=r, column=82).value == id):
                    if (sheet.cell(row=r, column=33).value != ""):
                        accounts[id][3].append(sheet.cell(row=r, column=33).value)
                    if (sheet.cell(row=r, column=75).value != ""):
                        accounts[id][3].append(sheet.cell(row=r, column=75).value)
                    if (sheet.cell(row=r, column=74).value != ""):
                        accounts[id][3].append(sheet.cell(row=r, column=74).value)
                    if (sheet.cell(row=r, column=78).value != ""):
                        accounts[id][3].append(sheet.cell(row=r, column=78).value)
                    if (sheet.cell(row=r, column=79).value != ""):
                        accounts[id][3].append(sheet.cell(row=r, column=79).value)

def writeEmail(account):
    s = ""
    if len(account) < 3:
        return s
    a = account[0].split(',')
    nm = ""
    if(len(a) > 1):
        nm = a[1]
    else:
        nm = a

    s = "Dear " + str(nm).strip() + ",\n\n"
    s += "We have, in our records, that you pledged "
    if(account[2] == 0):
        s += "a matanah, customarily $18, to our shul \nfor the "
    else:
        s += "to give $" + str(account[2]).strip() + ", to our shul \nfor the "

    if account[1] is None:
        account[1] = "aliya"
    s += account[1].replace('aliaya', 'aliya') + "."
    s += "  We very much appreciate your pledge!\n\nYou can pay online, or, if "
    s += "you prefer, you can send a check to the office at\n\nKadima Toras Moshe\n113 Washington "
    s += "Street\nBrighton, MA 02135\n\nIf this email is in error, please let us know so we can "
    s += "\ncorrect our records."
    s += "\n\nThanks so much!\n\nBest and Good Shabbos! "
    #print (s)
    return s

def sendTo(account):
    arr = []
    arr2 = []
    for email in account[3]:
        if(isinstance(email, unicode) == True):
            if (email.find("@") > 0):
                arr.append(email)

    for email in arr:
        if (email not in arr2):
            arr2.append(email)
    return arr2

def fire(fromaddr, toaddrs, msg):
    global server
    server.sendmail(fromaddr, toaddrs, msg)
    print("#"*45)
    print(msg)
    print("#"*45)

def sendItAll():
    global server
    global fromaddr
    global toaddrs

    usedAddressList = []
    for accounts in accountArray:
        for a in accounts:
            #print(len(accounts[a]))
            #print(accounts[a])
            msgTxt = writeEmail(accounts[a])
            list = sendTo(accounts[a])
            usedAddressList = []
            #if(len(list) < 1):
            #    print("Not Sent: " + str(a))
            for email in list:
                toaddrs = email
                if email not in usedAddressList:
                    msg = "\r\n".join([
                      "From: " + fromaddr,
                      "To: " + toaddrs,
                      "Subject: Aliyah Matanah",
                      "",
                      msgTxt
                      ])

                    fire(fromaddr, toaddrs, msg)
                    usedAddressList.append(email)

def printCounts():
    global notAliyah
    global publicLines
    global closedLines
    global paymentLines
    global aliyahCount

    print("#############################################################")
    print("### Counts                                                ###")
    print("### Lines that were not Aliyah: " + str(notAliyah) + "                ###")
    print("### Lines that were public: " + str(notAliyah) + "                    ###")
    print("### Lines that were closed: " + str(notAliyah) + "                    ###")
    print("### Lines that were payments: " + str(notAliyah) + "                  ###")
    print("### Lines that were Aliyahs: " + str(notAliyah) + "                   ###")
    print("### Total: " + str(totalLines) + "                                    ###")
    print("#############################################################")

#######################################################

def runProcess():
    global server
    trx = openTrx()
    ppl = openPPL()
    getTrx(trx[trx.sheetnames[0]])
    getEmail(ppl[ppl.sheetnames[0]])
    server = smtplib.SMTP(smtpvar)
    server.ehlo()
    server.starttls()
    server.login(username,password)
    sendItAll()
    server.quit()
    printCounts()

runWhatever(runProcess)
