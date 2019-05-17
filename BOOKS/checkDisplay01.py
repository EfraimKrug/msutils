#!/usr/bin/env python
# -*- coding: utf-8 -*-
import subprocess
import os
import sys
import winreg
#########################################################
# get parent directory...
#sys.path.append(os.getcwd())
#sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])
#######################################################
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.scrolledtext as tkst
#######################################################
import csv
import shutil
import requests

from datetime import datetime
from datetime import timedelta
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

from functools import partial

import smtplib
from Profile import *
from SearchNames import *
from checkDisplay02 import *
from checkDisplay03 import *
from checkDisplay04 import *
from errorDisplay import *
#from AlefBet import *
#from periodProcess import *
####################################################################################################
### This is where the entire application opens - with a list of all deposits tracked so far -
### in all workbooks in the directory
####################################################################################################
class checkDisplay01:
    def __init__(self, master):
        self.cashcheckSwitch = ''
        self.ds = dict()        # {check_name: [check_number, memo, check_date, arrival_date, check_amount, check_image],
        self.sdata = dict()     # sheet by sheet...
        self.pdata = dict()
        self.cdata = []     # cash
        self.searchObj = ''
        self.depositName = []

        self.master = master
        self.master.configure(bg="teal", pady=34, padx=17)
        self.master.geometry('400x300')
        self.master.title('Kadima Toras-Moshe Check Tracking')

        self.frame = tk.Frame(self.master, width=360, height=260)
        self.frame.configure(bg="teal", pady=2, padx=2)
        self.frame.grid(row=1, column=1)

        self.people = []
        self.pages = []
        self.files = []
        self.workbooks = dict()
        self.workingFile = 'DailyLog'

        #self.tkvar = ''
        self.EXCELEXE = ''

        self.dropInit = True
        self.runProcess()

    def getExcel(self):
        handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")

        num_values = winreg.QueryInfoKey(handle)[1]
        for i in range(num_values):
            for x in winreg.EnumValue(handle, i):
                if(str(x).find("EXCEL") > -1):
                    self.EXCELEXE = x

    def error_window(self, message):
        self.newWindow = tk.Toplevel(self.master)
        self.app = errorDisplay(self.newWindow, "Crash & Burn: " + message)

    def show_image(self, img):
        try:
            fileName = checkDir + img + ".pdf"
            path_to_pdf = os.path.abspath(fileName)
            path_to_acrobat = os.path.abspath(AcrobatPath)
            process = subprocess.Popen([path_to_acrobat, '/A', 'page=1', path_to_pdf], shell=False, stdout=subprocess.PIPE)
            process.wait()
        except:
            self.error_window("Sorry, that file can not be found!")

    def getFiles(self):
        self.files = []
        file_list=os.listdir(dailyLogDir)
        for  fileN in file_list:
            if fileN.find('xlsx') > 0:
                self.files.append(fileN[0:-5])

        return self.files

    def getCurrentWorkbook(self):
        return self.workbooks[self.workingFile]

    def openDailyLog(self):
        for file in self.files:
            self.workbooks[file] = load_workbook(dailyLogDir + '\\' + file + '.xlsx')
        return self.workbooks

    def createSheet(self):
        sheetNameNew = True
        dt = datetime.today().strftime('%B-%d')
        da = dt.split('-')
        sheetName = da[0] + str(da[1])

        #dailyLog = self.openDailyLog()
        for wb in self.workbooks:
            for name in self.workbooks[wb].sheetnames:
                if name == sheetName:
                    sheetNameNew = False

        # for name in dailyLog.sheetnames:
        #     if name == sheetName:
        #         sheetNameNew = False

        if not sheetNameNew:
            return

        newSheet = self.getCurrentWorkbook().create_sheet(title = sheetName)

        newSheet = self.getCurrentWorkbook()[sheetName]
        self.buildPage(newSheet)
        self.getCurrentWorkbook().save(dailyLogDir + '\\' + self.workingFile + '.xlsx')

    def openNewSheet(self):
        self.createSheet()
        self.getExcel()
        os.system("start  \"" + self.EXCELEXE + "\" \"" + dailyLogDir + "\\" + self.workingFile + ".xlsx\"")

    def shiftWBook(self):
        sheetNames = []
        lastThree = []
        dailyLog = self.openDailyLog()
        newBookName = ''

        for name in dailyLog.sheetnames:
            sheetNames.append(name)

        if len(sheetNames) > 4:
            newBookName = sheetNames[4][0:-2]

        if len(newBookName) < 3:
            return

        newFileName = dailyLogDir + "\\" + newBookName + '.xlsx'
        workingFile = dailyLogDir + "\\" + self.workingFile + '.xlsx'

        #for sName in sheetNames:
        # if datetime.today().day > 25:
        #     dt = datetime.today() + timedelta(days=10)
        #     newFileName = dailyLogDir + "\\" + dt.strftime('%B') + '.xlsx'
        #     print(newFileName)
        # else:
        #     return

        try:
            fh = open(newFileName, 'r')
            print("Sorry - we have already cycled the files")
            return
        except FileNotFoundError:
            print("Processing new file...")

        #    print(sName)

        for i in range(-3, 0):
            lastThree.append(sheetNames[i])

        #print("opening workbook")
        wb = Workbook()

        for sheet in lastThree:
            newSheet = wb.create_sheet(title = sheet)
            self.buildPage(newSheet)
            oldSheet = dailyLog[sheet]
            for colN in range(1,20):
                for rowN in range(1,35):
                    #print(oldSheet.cell(row=rowN, column=colN).value)
                    newSheet.cell(row=rowN, column=colN).value = oldSheet.cell(row=rowN, column=colN).value

        firstSheet = wb['Sheet']
        wb.remove_sheet(firstSheet)
        dailyLog.save(newFileName)
        wb.save(workingFile)


    def buildPage(self, newSheet):
        al = Alignment(horizontal='center', vertical='center')
        newSheet.cell(row=1,column=1).value = datetime.today().strftime('%d-%B-%Y')
        newSheet.cell(row=1,column=7).value = "Deposit: "
        newSheet.cell(row=1,column=7).alignment = al

        newSheet.cell(row=2,column=2).value = "Name on Check"
        newSheet.cell(row=2,column=2).alignment = al
        newSheet.cell(row=2,column=3).value = "Memo"
        newSheet.cell(row=2,column=3).alignment = al
        newSheet.cell(row=2,column=4).value = "Date on Check"
        newSheet.cell(row=2,column=4).alignment = al
        newSheet.cell(row=2,column=5).value = "Amount"
        newSheet.cell(row=2,column=5).alignment = al
        newSheet.cell(row=2,column=6).value = "Image"
        newSheet.cell(row=2,column=6).alignment = al

        newSheet.cell(row=3,column=1).value = "Cash"
        newSheet.cell(row=3,column=1).alignment = al
        newSheet.cell(row=11,column=4).value = "Total: "
        newSheet.cell(row=11,column=5).value = "=SUM(E4:E10)"

        newSheet.cell(row=12,column=1).value = "Check No."
        newSheet.cell(row=12,column=1).alignment = al
        newSheet.cell(row=31,column=3).value = "Sub Total: "
        newSheet.cell(row=31,column=5).value = "=SUM(E13:E30)"
        newSheet.cell(row=32,column=3).value = "Grand Total: "
        newSheet.cell(row=32,column=5).value = "=SUM(E11,E31)"

        #newSheet.column_dimensions[0].width = 20.71
        newSheet.column_dimensions['A'].width = 20
        newSheet.column_dimensions['B'].width = 33
        newSheet.column_dimensions['C'].width = 51
        newSheet.column_dimensions['D'].width = 23
        newSheet.column_dimensions['E'].width = 12
        newSheet.column_dimensions['F'].width = 19
        newSheet.column_dimensions['G'].width = 11

    def parseName(self, name):
        day = name[-2:]
        month = name[0:-2]
        return (day, month)

    def loadRow(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]


        name = month+day

        dName = ''
        for depArr in self.depositName:
            if name == depArr[1]:
                dName = depArr[0]

        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  dName,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        if not peep in self.people:
            self.people.append(peep)

        arr.append(newRow)
        self.ds[sheet.cell(row=current_row, column=2).value] = arr
        if not name in self.sdata:
            self.sdata[name] = dict()

        self.sdata[name][sheet.cell(row=current_row, column=2).value] = arr
        if not peep in self.pdata:
            self.pdata[peep] = dict()

        self.pdata[peep][sheet.cell(row=current_row, column=2).value] = arr


    def loadRowCash(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]

        name = month+day

        dName = ''
        for depArr in self.depositName:
            if name == depArr[1]:
                dName = depArr[0]

        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  dName,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        self.cdata.append(newRow)

    def getSheet(self, name, sheet, wb):
        (day, month) = self.parseName(name)
        if not name in self.pages:
            self.pages.append([name, wb])

        self.depositName.append([str(sheet.cell(row=2,column=7).value), name, wb])

        for r in range(3, sheet.max_row):
            if(str(sheet.cell(row=r,column=1).value).lower() == 'cash'):
                self.cashcheckSwitch = 'cash'
            if(str(sheet.cell(row=r,column=1).value).lower().find('check') > -1):
                self.cashcheckSwitch = 'check'

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('check') > -1):
                self.loadRow(month, day, sheet, r)

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('cash') > -1):
                self.loadRowCash(month, day, sheet, r)


# on change dropdown value
    # def change_dropdown(self, *args):
    #     self.workingFile = self.tkvar.get()


    def change_dropdown2(self, *args):
        if self.dropInit:
            self.dropInit = False
            return
        wb = ""
        name = self.tkvar2.get()
        for a in self.pages:
            if a[0] == name:
                wb = a[1]

        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay02(self.newWindow, name, wb)

    def showCash(self):
        total = 0
        for line in self.cdata:
            total += line[5]
        total = "{:.2f}".format(float(total))
        total = "Cash: $" + total
        self.label07 = tk.Label(self.frame, text=total, bg="teal", fg="yellow", font='Helvetica 10 bold')
        self.label07.grid(row=1, column=18, padx=4, pady=4, sticky=tk.NW)

    def showPerson(self, name, args):
        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay03(self.newWindow, name)

    def showDeposit(self, name, args):
        wb = ""
        dName = ""
        sName = ""
        for depArr in self.depositName:
            if name == depArr[0]:
                dName = depArr[0]
                sName = depArr[1]
                wb = depArr[2]

        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay04(self.newWindow, dName, sName, wb)

    # link function to change change_dropdown
    def showData(self):
        total = 0
        self.label03 = []
        #self.button01 = []

        self.peepFunctions = []
        self.deposits = []

        fileNames = []

        fileList = self.getFiles()
        # self.tkvar = tk.StringVar(self.master)
        # self.tkvar.trace('w', self.change_dropdown)
        # self.tkvar.set(fileList[0]) # set the default option
        #
        self.tkvar2 = tk.StringVar(self.master)
        self.tkvar2.trace('w', self.change_dropdown2)
        self.tkvar2.set(self.pages[0][0]) # set the default option

        # pagesPopup = tk.OptionMenu(self.frame, self.tkvar, *fileList)
        # pagesPopup.grid(row = 1, column =6, padx=10, pady=10, sticky=tk.EW)
        pages = []
        for a in self.pages:
            pages.append(a[0])

        pagesPopup2 = tk.OptionMenu(self.frame, self.tkvar2, *pages)
        pagesPopup2.grid(row = 1, column =5, padx=10, pady=10, sticky=tk.EW)

        self.button01 = tk.Button(self.frame, text="Shift", command=partial(self.shiftWBook))
        self.button01.grid(row=1, column=2, columnspan=1, padx=10, pady=10, sticky=tk.EW)

        self.button02 = tk.Button(self.frame, text="Open Excel", command=partial(self.openNewSheet))
        self.button02.grid(row=1, column=3, columnspan=1, padx=10, pady=10, sticky=tk.EW)

        row_num = 6
        self.headline03 = tk.Label(self.frame, text=" Deposits ", bg="teal", fg="yellow")
        self.headline03.grid(row=3, column=3, padx=4, pady=2, sticky=tk.W)

        for e1 in self.ds:
            for e2 in self.ds[e1]:
                if not e2[1] in self.deposits:
                    self.deposits.append(e2[1])

        for d in self.deposits:
            self.label03.append(tk.Label(self.frame, text=d, bg="teal", fg="yellow"))
            self.label03[len(self.label03)-1].grid(row=row_num, column=3, padx=4, pady=4, sticky=tk.NW)
            self.label03[len(self.label03)-1].bind("<Button-1>", partial(self.showDeposit, d))
            row_num = row_num + 1

    def sheetExists(self, sheet):
        dailyLog = self.openDailyLog()
        for name in dailyLog.sheetnames:
            if name == sheet:
                return True
        return False

    def runProcess(self):
        self.getFiles()
        self.openDailyLog()
        for wb in self.workbooks:
            for name in self.workbooks[wb].sheetnames:
                self.total = 0
                self.getSheet(name, self.workbooks[wb][name], wb)
        self.showData()
