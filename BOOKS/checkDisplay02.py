#!/usr/bin/env python
# -*- coding: utf-8 -*-
import subprocess
import os
import sys
#########################################################
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.scrolledtext as tkst
#######################################################
import csv
import shutil
import requests

from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side
from functools import partial

import smtplib
from Profile import *
from checkDisplay03 import *
from checkDisplay04 import *
####################################################################################################
class checkDisplay02:
    def __init__(self, master, oSheetName):
        self.cashcheckSwitch = ''
        self.ds = dict()
        self.oSheetName = oSheetName
        self.sdata = dict()     # sheet by sheet...
        self.pdata = dict()
        self.cdata = []
        self.depositName = ''

        self.master = master
        self.master.configure(bg="teal", pady=34, padx=17)
        self.master.geometry('700x700')
        self.master.title('Edit Your Display Data')

        self.frame = tk.Frame(self.master, width=460, height=360)
        self.frame.configure(bg="teal", pady=2, padx=2)
        self.frame.grid(row=1, column=1)

        self.people = []
        self.pages = []
        self.tkvar = ''

        self.runProcess(self.oSheetName)

    def getEachKeyStroke(self, key):
        print("here")

    def setOutlabel(self, newLabel):
        self.OutLabel['text'] = newLabel

    def upArrow(self, key):
        print("here")

    def downArrow(self, key):
        print("here")

    def pre_new_window(self, key):
        self.new_window()

    def new_empty_window(self):
        self.newWindow = tk.Toplevel(self.master)

    def new_config_window(self):
        self.newWindow = tk.Toplevel(self.master)

    def show_image(self, img):
        fileName = "C:\\Users\\KTM\\Documents\\EMK\\BOOKS\\Checks\\" + img + ".pdf"
        path_to_pdf = os.path.abspath(fileName)
        path_to_acrobat = os.path.abspath('C:\\Program Files (x86)\\Adobe\\Acrobat Reader DC\\Reader\\AcroRd32.exe')
        process = subprocess.Popen([path_to_acrobat, '/A', 'page=1', path_to_pdf], shell=False, stdout=subprocess.PIPE)
        process.wait()

    def openDailyLog(self):
        wb = load_workbook(dailyLogDir + '\\dailyLog.xlsx')
        return wb

    def parseName(self, name):
        day = name[-2:]
        month = name[0:-2]
        return (day, month)

    def loadRowCash(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]

        name = month+day
        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  self.depositName,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        # if not peep in self.people:
        #     self.people.append(peep)

        self.cdata.append(newRow)
        #print(arr)
        # self.ds[sheet.cell(row=current_row, column=2).value] = arr
        # if not name in self.sdata:
        #     self.sdata[name] = dict()
        #
        # self.sdata[name][sheet.cell(row=current_row, column=2).value] = arr
        # if not peep in self.pdata:
        #     self.pdata[peep] = dict()
        #
        # self.pdata[peep][sheet.cell(row=current_row, column=2).value] = arr


    def loadRow(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]

        name = month+day
        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  self.depositName,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        if not peep in self.people:
            self.people.append(peep)

        arr.append(newRow)
        self.ds[sheet.cell(row=current_row, column=2).value] = arr
        if not name in self.sdata:
            self.sdata[name] = dict()

        self.sdata[name][sheet.cell(row=current_row, column=2).value] = arr
        if not peep in self.pdata:
            self.pdata[peep] = dict()

        self.pdata[peep][sheet.cell(row=current_row, column=2).value] = arr


    def getSheet(self, name, sheet):
        (day, month) = self.parseName(name)
        if not name in self.pages:
            self.pages.append(name)

        self.depositName = str(sheet.cell(row=2,column=7).value)

        for r in range(3, sheet.max_row):
            if(str(sheet.cell(row=r,column=1).value).lower() == 'cash'):
                self.cashcheckSwitch = 'cash'
            if(str(sheet.cell(row=r,column=1).value).lower().find('check') > -1):
                self.cashcheckSwitch = 'check'

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('check') > -1):
                self.loadRow(month, day, sheet, r)

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('cash') > -1):
                self.loadRowCash(month, day, sheet, r)


    def showCash(self):
        total = 0
        for line in self.cdata:
            total += line[4]
        total = "{:.2f}".format(float(total))
        total = "Cash: $" + total
        self.label07 = tk.Label(self.frame, text=total, bg="teal", fg="yellow", font='Helvetica 10 bold')
        self.label07.grid(row=1, column=18, padx=4, pady=4, sticky=tk.NW)

    def showPerson(self, name, args):
        #print(name)
        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay03(self.newWindow, name)

    def showDeposit(self, name, args):
        #print(name)
        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay04(self.newWindow, name)

# on change dropdown value
    # def change_dropdown(self, *args):
    #     print( self.tkvar.get() )
    #
    # def change_dropdown2(self, *args):
    #     #self.ds = self.sdata[ self.tkvar2.get() ]
    #     #self.frame = None
    #     #self.showData()
    #     print( self.tkvar2.get() )
    #
    # link function to change change_dropdown
    def showData(self):
        #frame = tk.Frame(self.master, width=700, height=300)
        total = 0
        self.label01 = []
        self.label02 = []
        self.label03 = []
        self.label04 = []
        self.label05 = []
        self.label06 = []
        self.button01 = []
        fileNames = []

        row_num = 6
        self.title = tk.Label(self.frame, text=self.oSheetName, bg="teal", fg="yellow", font='Helvetica 10 bold')
        self.title.grid(row=1, column=1, padx=4, pady=4, sticky=tk.NW)

        self.headline01 = tk.Label(self.frame, text="Name", bg="teal", fg="yellow")
        self.headline01.grid(row=3, column=2, padx=4, pady=2, sticky=tk.W)

        self.headline02 = tk.Label(self.frame, text=" Date ", bg="teal", fg="yellow")
        self.headline02.grid(row=3, column=4, padx=4, pady=2, sticky=tk.W)

        self.headline03 = tk.Label(self.frame, text=" Deposit ", bg="teal", fg="yellow")
        self.headline03.grid(row=3, column=6, padx=4, pady=2, sticky=tk.W)

        self.headline04 = tk.Label(self.frame, text="Check #", bg="teal", fg="yellow")
        self.headline04.grid(row=3, column=8, padx=4, pady=2, sticky=tk.W)

        self.headline05 = tk.Label(self.frame, text="Amount", bg="teal", fg="yellow")
        self.headline05.grid(row=3, column=10, padx=4, pady=2, sticky=tk.W)

        #self.headline05 = tk.Label(self.frame, text="Sheet", bg="teal", fg="yellow")
        #self.headline05.grid(row=1, column=10, padx=4, pady=2, sticky=tk.W)

        self.headline06 = tk.Label(self.frame, text="Sheet Total", bg="teal", fg="yellow")
        self.headline06.grid(row=3, column=14, padx=4, pady=2, sticky=tk.W)

        self.headline07 = tk.Label(self.frame, text="Image", bg="teal", fg="yellow")
        self.headline07.grid(row=3, column=16, padx=4, pady=2, sticky=tk.W)


        sortedKeys = []
        for key in self.ds:
            sortedKeys.append(key)
        sortedKeys.sort()
        lastName = ''
        pEnt = ''
        totals = dict()

        self.showCash()

        for ent in sortedKeys:
            for e in self.ds[ent]:
                pEnt = ent
                if lastName == ent:
                    pEnt = ''
                lastName = ent
                self.label01.append(tk.Label(self.frame, text=pEnt, bg="teal", fg="yellow"))
                self.label01[len(self.label01)-1].grid(row=row_num, column=2, padx=4, pady=4, sticky=tk.NW)
                self.label01[len(self.label01)-1].bind("<Button-1>", partial(self.showPerson, pEnt))

                self.label02.append(tk.Label(self.frame, text=e[3], bg="teal", fg="yellow"))
                self.label02[len(self.label02)-1].grid(row=row_num, column=4, padx=4, pady=4, sticky=tk.NW)

                self.label03.append(tk.Label(self.frame, text=e[1], bg="teal", fg="yellow"))
                self.label03[len(self.label03)-1].grid(row=row_num, column=6, padx=4, pady=4, sticky=tk.NW)
                self.label03[len(self.label03)-1].bind("<Button-1>", partial(self.showDeposit, e[1]))

                self.label04.append(tk.Label(self.frame, text=e[0], bg="teal", fg="yellow"))
                self.label04[len(self.label03)-1].grid(row=row_num, column=8, padx=4, pady=4, sticky=tk.NW)
                #self.label03[len(self.label03)-1].bind("<Button-1>", partial(self.showDeposit, e[1]))

                fAmt = "{:.2f}".format(float(e[5]))
                self.label05.append(tk.Label(self.frame, text=fAmt, bg="teal", fg="yellow"))
                self.label05[len(self.label04)-1].grid(row=row_num, column=10, padx=4, pady=4, sticky=tk.NW)

                #self.label05.append(tk.Label(self.frame, text=e[6], bg="teal", fg="yellow"))
                #self.label05[len(self.label05)-1].grid(row=row_num, column=10, padx=4, pady=4, sticky=tk.NW)

                if e[7] in totals:
                    totals[e[7]] = totals[e[7]] + e[5]
                else:
                    totals[e[7]] = e[5]

                fAmt2 = "{:.2f}".format(float(totals[e[7]]))
                self.label06.append(tk.Label(self.frame, text="$" + str(fAmt2), bg="teal", fg="yellow"))
                self.label06[len(self.label06)-1].grid(row=row_num, column=14, padx=4, pady=4, sticky=tk.NW)

                self.button01.append(tk.Button(self.frame, text="View", command=partial(self.show_image, e[6])))
                self.button01[len(self.button01)-1].grid(row=row_num, column=16, columnspan=2, padx=4, pady=4, sticky=tk.EW)
                total = total + e[5]
                row_num += 1
                line = ''

    def sheetExists(self, sheet):
        dailyLog = self.openDailyLog()
        for name in dailyLog.sheetnames:
            if name == sheet:
                return True
        return False

    def runProcess(self, name):
        dailyLog = self.openDailyLog()
        #for name in dailyLog.sheetnames:
        self.total = 0
        self.getSheet(name, dailyLog[name])
        self.showData()
        #print(self.pdata)
