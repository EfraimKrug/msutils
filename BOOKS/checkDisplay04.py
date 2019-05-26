#!/usr/bin/env python
# -*- coding: utf-8 -*-
import subprocess
import os
import sys
#########################################################
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.scrolledtext as tkst
import tkinter.font as tkFont
#######################################################
import csv
import shutil
import requests

from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side
from functools import partial

import smtplib
from Profile import *
from checkDisplay02 import *
from errorDisplay import *
####################################################################################################
### Showing the individual deposit - that will probably span across spreadsheets -
####################################################################################################
class checkDisplay04:
    def __init__(self, master, depositName, sheet, wb):
        self.cashcheckSwitch = ''
        self.ds = dict()
        self.depositName = depositName
        self.sheetName = sheet
        self.wb = wb
        self.sdata = dict()     # sheet by sheet...
        self.pdata = dict()
        self.cdata = []

        self.master = master
        self.fullHeight = self.master.winfo_screenheight()
        self.fullWidth = self.master.winfo_screenwidth()
        #self.master.configure(bg="teal", pady=34, padx=17)
        self.master.configure(bg="teal", pady=3, padx=1)
        self.master.geometry('530x700')
        #self.master.geometry('700x%s' % (self.fullHeight))
        self.master.title(depositName)


        self.frame = tk.Frame(self.master, width=310, height=self.fullHeight)
        #self.frame = tk.Frame(self.master, width=460, height=360)
        self.frame.configure(bg="teal", pady=2, padx=2)
        self.frame.grid(row=1, column=1)

        self.people = []
        self.pages = []
        self.tkvar = ''
        self.checkTotal = 0
        self.cashTotal = 0

        self.depositWB = ""
        self.newDepositSheet = ""

        self.runProcess(self.depositName)

    # def setOutlabel(self, newLabel):
    #     self.OutLabel['text'] = newLabel
    #
    # def pre_new_window(self, key):
    #     self.new_window()
    #
    # def new_empty_window(self):
    #     self.newWindow = tk.Toplevel(self.master)
    #
    def error_window(self, message):
        self.newWindow = tk.Toplevel(self.master)
        self.app = errorDisplay(self.newWindow, "Crash & Burn: " + message)

    def show_image(self, img):
        try:
            fileName = checkDir + img + ".pdf"
            path_to_pdf = os.path.abspath(fileName)
            path_to_acrobat = os.path.abspath(AcrobatPath)
            process = subprocess.Popen([path_to_acrobat, '/A', 'page=1', path_to_pdf], shell=False, stdout=subprocess.PIPE)
            process.wait()
        except:
            self.error_window("Sorry, that file can not be found!")

    def openDailyLog(self):
        wb = load_workbook(dailyLogDir + '\\' + self.wb + '.xlsx')
        return wb

    def parseName(self, name):
        day = name[-2:]
        month = name[0:-2]
        return (day, month)

    def loadRowCash(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]

        name = month+day
        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        self.cdata.append(newRow)


    def loadRow(self, month, day, sheet, current_row):
        arr = []
        if(sheet.cell(row=current_row, column=2).value in self.ds):
            arr = self.ds[sheet.cell(row=current_row, column=2).value]

        name = month+day
        peep = sheet.cell(row=current_row, column=2).value

        newRow = [sheet.cell(row=current_row, column=1).value,
                  sheet.cell(row=current_row, column=3).value,
                  str(sheet.cell(row=current_row, column=4).value)[0:10],
                  str(month) + "-" + str(day),
                  sheet.cell(row=current_row, column=5).value,
                  sheet.cell(row=current_row, column=6).value,
                  name]

        if not peep in self.people:
            self.people.append(peep)

        arr.append(newRow)
        self.ds[sheet.cell(row=current_row, column=2).value] = arr
        if not name in self.sdata:
            self.sdata[name] = dict()

        self.sdata[name][sheet.cell(row=current_row, column=2).value] = arr
        if not peep in self.pdata:
            self.pdata[peep] = dict()

        self.pdata[peep][sheet.cell(row=current_row, column=2).value] = arr


    def getSheet(self, name, sheet, depositName):

        if(not sheet.cell(row=2, column=7).value == depositName):
            return False

        (day, month) = self.parseName(name)
        if not name in self.pages:
            self.pages.append(name)

        for r in range(3, sheet.max_row):
            if(str(sheet.cell(row=r,column=1).value).lower() == 'cash'):
                self.cashcheckSwitch = 'cash'
            if(str(sheet.cell(row=r,column=1).value).lower().find('check') > -1):
                self.cashcheckSwitch = 'check'

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('check') > -1):
                self.loadRow(month, day, sheet, r)

            if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('cash') > -1):
                self.loadRowCash(month, day, sheet, r)

        return True

    def showPerson(self, name, args):
        self.newWindow = tk.Toplevel(self.master)
        self.app = checkDisplay03(self.newWindow, name)

    def showCash(self):
        total = 0
        for line in self.cdata:
            total += line[4]

        self.cashTotal = total
        total = "{:.2f}".format(float(total))
        total = "Cash: $" + total
        self.label07 = tk.Label(self.frame, text=total, bg="teal", fg="yellow", font='Helvetica 10 bold')
        self.label07.grid(row=1, column=18, padx=4, pady=4, sticky=tk.NW)

# on change dropdown value
    # def change_dropdown(self, *args):
    #     print( self.tkvar.get() )
    #
    # def change_dropdown2(self, *args):
    #     #self.ds = self.sdata[ self.tkvar2.get() ]
    #     #self.frame = None
    #     #self.showData()
    #     print( self.tkvar2.get() )
    #
    # link function to change change_dropdown
    # def buildNewDeposit(self, sheet):
    #         for r in range(3, sheet.max_row):
    #             if(str(sheet.cell(row=r,column=1).value).lower() == 'cash'):
    #                 self.cashcheckSwitch = 'cash'
    #             if(str(sheet.cell(row=r,column=1).value).lower().find('check') > -1):
    #                 self.cashcheckSwitch = 'check'
    #
    #             if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('check') > -1):
    #                 self.loadRow(month, day, sheet, r)
    #
    #             if(sheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('cash') > -1):
    #                 self.loadRowCash(month, day, sheet, r)
    #
    def openDepositWorkBook(self, depositName):
        self.depositWB = load_workbook(depositDir + '\\Deposits.xlsx')
        for name in self.depositWB.sheetnames:
            if name == depositName:
                #already created the deposit!
                return False
        #print ("nope")
        self.newDepositSheet = self.depositWB.create_sheet(title = depositName)
        return True

    def copyData(self, oldSheet, newSheet):
            newCheckRow = 3
            newCashRow = 50

            newSheet.cell(row=1, column=2).value = "Name on check"
            newSheet.cell(row=1, column=3).value = "Memo"
            newSheet.cell(row=1, column=4).value = "Date on check"
            newSheet.cell(row=1, column=5).value = "Amount"
            newSheet.cell(row=1, column=6).value = "Image"
            newSheet.cell(row=2, column=1).value = "Check No."
            newSheet.cell(row=49, column=1).value = "Cash"
            #print("==>" + oldSheet.cell(row=2, column=7).value)
            #self.newDepositSheet.cell(row=5,column=1).value = "HERE"
            #return
            for r in range(3, oldSheet.max_row):
                if(str(oldSheet.cell(row=r,column=1).value).lower() == 'cash'):
                    self.cashcheckSwitch = 'cash'
                if(str(oldSheet.cell(row=r,column=1).value).lower().find('check') > -1):
                    self.cashcheckSwitch = 'check'

                if(oldSheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('check') > -1):
                    for c in range(1, 8):
                        newSheet.cell(row=newCheckRow, column=c).value = oldSheet.cell(row=r,column=c).value
                    newCheckRow += 1

                if(oldSheet.cell(row=r, column=2).value and self.cashcheckSwitch.find('cash') > -1):
                    for c in range(1, 8):
                        newSheet.cell(row=newCashRow, column=c).value = oldSheet.cell(row=r,column=c).value
                    newCashRow += 1

    def makeDeposit(self, name, args):
        if(self.openDepositWorkBook(name)):
            dailyLog = self.openDailyLog()
            for sheetName in dailyLog.sheetnames:
                if(name == dailyLog[sheetName].cell(row=2,column=7).value):
                    #print(dailyLog[sheetName].cell(row=2,column=7).value)
                    self.copyData(dailyLog[sheetName], self.newDepositSheet)

        self.depositWB.save(depositDir + '\\Deposits.xlsx')

                    #print(dailyLog[sheetName].cell(row=2,column=7).value + "::" + sheetName)

    def showData(self, rNum):
        #frame = tk.Frame(self.master, width=700, height=300)
        #for x in self.ds:
        #    print(x)
        row_num = rNum
        # total = 0
        self.label01 = []
        self.label02 = []
        self.label03 = []
        self.label04 = []
        self.label05 = []
        self.label06 = []
        self.button01 = []
        fileNames = []

        self.title = tk.Label(self.frame, text=self.depositName, bg="teal", fg="yellow", font='Helvetica 10 bold')
        self.title.grid(row=1, column=1, padx=1, pady=4, sticky=tk.NW)
        self.title.bind("<Button-1>", partial(self.makeDeposit, self.depositName))

        self.headline01 = tk.Label(self.frame, text="Name", bg="teal", fg="yellow")
        self.headline01.grid(row=3, column=2, padx=1, pady=2, sticky=tk.W)

        self.headline02 = tk.Label(self.frame, text=" Date ", bg="teal", fg="yellow")
        self.headline02.grid(row=3, column=4, padx=1, pady=2, sticky=tk.W)

        self.headline03 = tk.Label(self.frame, text="Check #", bg="teal", fg="yellow")
        self.headline03.grid(row=3, column=6, padx=1, pady=2, sticky=tk.W)

        self.headline04 = tk.Label(self.frame, text="Amount", bg="teal", fg="yellow")
        self.headline04.grid(row=3, column=8, padx=1, pady=2, sticky=tk.W)

        #self.headline05 = tk.Label(self.frame, text="Sheet", bg="teal", fg="yellow")
        #self.headline05.grid(row=1, column=10, padx=4, pady=2, sticky=tk.W)

        # self.headline06 = tk.Label(self.frame, text="Sheet Total", bg="teal", fg="yellow")
        # self.headline06.grid(row=3, column=12, padx=4, pady=2, sticky=tk.W)

        self.headline07 = tk.Label(self.frame, text="Image", bg="teal", fg="yellow")
        self.headline07.grid(row=3, column=14, padx=1, pady=2, sticky=tk.W)


        sortedKeys = []
        for key in self.ds:
            sortedKeys.append(key)
        sortedKeys.sort()
        lastName = ''
        pEnt = ''
        # totals = dict()
        self.master.geometry('530x700')
        if len(sortedKeys) > 15:
            self.master.geometry('530x%s' % (self.fullHeight))

        self.showCash()
        label_font = tkFont.Font(family='Arial', size=8)
        #row_num = 6
        for ent in sortedKeys:
            for e in self.ds[ent]:
                pEnt = ent
                if lastName == ent:
                    pEnt = ''
                lastName = ent
                self.label01.append(tk.Label(self.frame, text=pEnt, font=label_font, bg="teal", fg="yellow"))
                self.label01[len(self.label01)-1].grid(row=row_num, column=2, padx=1, pady=1, sticky=tk.NW)
                self.label01[len(self.label01)-1].bind("<Button-1>", partial(self.showPerson, pEnt))

                self.label02.append(tk.Label(self.frame, text=e[3], font=label_font, bg="teal", fg="yellow"))
                self.label02[len(self.label02)-1].grid(row=row_num, column=4, padx=1, pady=1, sticky=tk.NW)

                self.label03.append(tk.Label(self.frame, text=e[0], font=label_font, bg="teal", fg="yellow"))
                self.label03[len(self.label03)-1].grid(row=row_num, column=6, padx=1, pady=1, sticky=tk.NW)

                fAmt = "{:.2f}".format(float(e[4]))
                self.label05.append(tk.Label(self.frame, text=fAmt, font=label_font, bg="teal", fg="yellow"))
                self.label05[len(self.label04)-1].grid(row=row_num, column=8, padx=1, pady=1, sticky=tk.NW)

                #self.label05.append(tk.Label(self.frame, text=e[6], bg="teal", fg="yellow"))
                #self.label05[len(self.label05)-1].grid(row=row_num, column=10, padx=4, pady=4, sticky=tk.NW)

                # if e[6] in totals:
                #     totals[e[6]] = totals[e[6]] + e[4]
                # else:
                #     totals[e[6]] = e[4]

                self.checkTotal = self.checkTotal + e[4]
                # fAmt2 = "{:.2f}".format(float(totals[e[6]]))
                # self.label06.append(tk.Label(self.frame, text="$" + str(fAmt2), bg="teal", fg="yellow"))
                # self.label06[len(self.label06)-1].grid(row=row_num, column=12, padx=4, pady=4, sticky=tk.NW)

                self.button01.append(tk.Button(self.frame, text="View", font=label_font, command=partial(self.show_image, e[5]),height=1))
                self.button01[len(self.button01)-1].grid(row=row_num, column=14, columnspan=2, padx=1, pady=1, sticky=tk.EW)
                # total = total + e[4]
                row_num += 1
                line = ''
        return row_num

    def sheetExists(self, sheet):
        dailyLog = self.openDailyLog()
        for name in dailyLog.sheetnames:
            if name == sheet:
                return True
        return False

    def runProcess(self, depositName):
        dailyLog = self.openDailyLog()
        self.total = 0
        row_num = 6
        for name in dailyLog.sheetnames:
            self.getSheet(name, dailyLog[name], depositName)

        row_num = self.showData(row_num)
        row_num = row_num + 3

        # total = "Cash: $" + total
        # self.label07 = tk.Label(self.frame, text=total, bg="teal", fg="yellow", font='Helvetica 10 bold')
        # self.label07.grid(row=1, column=18, padx=4, pady=4, sticky=tk.NW)

        fAmt = "Checks: $" + "{:.2f}".format(float(self.checkTotal))
        fAmtLabel = tk.Label(self.frame, text=fAmt, bg="teal", fg="yellow", font='Helvetica 10 bold')
        fAmtLabel.grid(row=row_num, column=18, padx=4, pady=4, sticky=tk.NW)

        row_num = row_num + 1
        fAmt = "Total: $" + "{:.2f}".format(float(self.checkTotal + self.cashTotal))
        fAmtLabel2 = tk.Label(self.frame, text=fAmt, bg="teal", fg="yellow", font='Helvetica 10 bold')
        fAmtLabel2.grid(row=row_num, column=18, padx=4, pady=4, sticky=tk.NW)

        #print(self.pdata)
