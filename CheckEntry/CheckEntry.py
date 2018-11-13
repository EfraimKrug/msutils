
#CheckProcess
import sys
import webbrowser
import csv
from openpyxl import load_workbook
import winreg
from profile import *

ACCOUNTS = dict()
CHECKS = dict()
CHROME_PATH = 'C:/Users/KTM/AppData/Local/Google/Chrome/Application/chrome.exe %s'
URLS = []

wDir = basedir + "\\CheckEntry"
def openwb():
  wb = load_workbook(wDir + "\\Accounts.xlsx")
  return wb

def getChrome():
    global CHROME_PATH
    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe")

    num_values = winreg.QueryInfoKey(handle)[1]
    for i in range(num_values):
        for x in winreg.EnumValue(handle, i):
            if(str(x).find("CHROME") > -1):
                CHROME_PATH = x + " %s"


def countOccurrences(sheet, _key):
    count = 0
    for r in range(1, sheet.max_row+1):
        _k = sheet.cell(row=r,column=1).value[0:str(sheet.cell(row=r,column=1).value).find(',')]
        if(_k.find(_key) == 0):
            count = count + 1
    return count

def loadACCOUNTS(wbook):
    _key = ""
    _key2 = ""
    _val = ""

    sheet = wbook[wbook.sheetnames[0]]
    for r in range(1, sheet.max_row+1):
        _key = sheet.cell(row=r,column=1).value[0:str(sheet.cell(row=r,column=1).value).find(',')]
        init = str(sheet.cell(row=r,column=1).value).find(',') + 1
        _key2 = sheet.cell(row=r,column=1).value[init:]
        _key2 = _key2.replace(' ','')

        if(countOccurrences(sheet, _key) > 1):
            _key = _key + _key2
            _key = _key.replace(',','')

        ACCOUNTS[_key] = sheet.cell(row=r,column=2).value

def loadURLS(checks):
    _key = ""
    _key2 = ""
    _val = ""

    sheet = checks[checks.sheetnames[0]]
    for r in range(1, sheet.max_row+1):
        init = str(sheet.cell(row=r,column=1).value).find(',') + 1
        if init < 1:
            _key = sheet.cell(row=r,column=1).value
        else:
            _key = sheet.cell(row=r,column=1).value[0:init-1]
        _key2 = sheet.cell(row=r,column=1).value[init:]
        _key2 = _key2.replace(' ','')

        print("CHECKING: " + _key)
        if _key in ACCOUNTS:
            ID = ACCOUNTS[_key]
            URLS.append("https://" + baseURL + "/admin/transaction_add.php?account_id=" + str(ID) + "&return=account")
        else:
            _key = _key + _key2
            _key = _key.replace(',','')
            #print("ELSING: " + _key + "::" + _key2)
            for _K in ACCOUNTS:
                if _K.find(_key) == 0:
                    #print("Found: " + _K)
                    ID = ACCOUNTS[_K]
                    URLS.append("https://" + baseURL + "/admin/transaction_add.php?account_id=" + str(ID) + "&return=account")

            #if _key in ACCOUNTS:
            #    ID = ACCOUNTS[_key]
            #    URLS.append("https://" + baseURL + "/admin/transaction_add.php?account_id=" + str(ID) + "&return=account")

def printURLS():
    for line in URLS:
        print(line)

def processURLS():
    for url in URLS:
        print(url)
        webbrowser.get(CHROME_PATH).open(url)
        print("Press enter to continue: ")
        x = raw_input()

def openChecksWB():
  wb = load_workbook(wDir + "\\Checks.xlsx")
  return wb

def printFile(wbook):
    sheet = wbook[wbook.sheetnames[0]]
    for r in range(1, sheet.max_row+1):
        print(sheet.cell(row=r,column=1).value + " :: " + str(sheet.cell(row=r,column=2).value))

accounts = openwb()
checks = openChecksWB()

getChrome()
print (CHROME_PATH)

loadACCOUNTS(accounts)
loadURLS(checks)
processURLS()
