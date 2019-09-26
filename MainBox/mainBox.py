from datetime import *
from tkinter import *
from tkinter import filedialog
import tkMessageBox
import subprocess
import os
import winreg

import time
from os import walk
import sys
import csv
import openpyxl
from openpyxl import Workbook
from functools import partial
#########################################################
# get parent directory...
#########################################################

sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import shutil
from Profile import *
# from CDCommonCode import *
# from BOOKS.checkMove01 import checkMove01

root = Tk()
# bgImage = PhotoImage(file=".\\MainBox\\KTMGate.PNG")
root.geometry('400x400')
root.title('Kadima Toras-Moshe System')

#fileSwitches
noYahrzeitFileSw = False
noYahrzeitFIXSw = False
noTransactionFileSw = False
noTransactionFIXSw = False

labelMessage = ""
EXCELEXE = ""
############################################
# windows: find the excel program
############################################
def getExcel():
    global EXCELEXE
    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")

    num_values = winreg.QueryInfoKey(handle)[1]
    for i in range(num_values):
        for x in winreg.EnumValue(handle, i):
            if(str(x).find("EXCEL") > -1):
                EXCELEXE = x

def openFile(fileName):
    getExcel()
    if fileName.find('Dep') > -1:
        os.system("start  \"" + EXCELEXE + "\" \"" + depositDir + "\\" + fileName + ".xlsx\"")
    if fileName.find('DailyLog') > -1:
        os.system("start  \"" + EXCELEXE + "\" \"" +  dailyLogDir + "\\" + fileName + ".xlsx\"")

# works
def refreshFile(fileName):
    wb = openpyxl.Workbook()
    ws = wb.active

    if 'yahrzeits' in fileName:
        newFileName = basedir + r'\\shulCloud\\yahrzeits.xlsx'
    if 'transactions' in fileName:
        newFileName = basedir + r'\\shulCloud\\transactions.xlsx'


    with open(fileName) as f:
        reader = csv.reader(f, delimiter=':')
        for row in reader:
            ws.append(row)

    try:
        os.remove(newFileName)
    except:
        print('no file to remove: ' + newFileName)

    wb.save(newFileName)

def checkMove():
    # cM1 = checkMove01()
    subprocess.call([batdir + '\\checkMove.bat'], shell=False)

def getNames():
    subprocess.call([batdir + '\\getNames.bat'], shell=False)
# works
def runYahr():
    subprocess.call([batdir + '\\yahr.bat'], shell=False)

def openBooks():
    # print(batdir)
    subprocess.call([batdir + '\\openBooks.bat', 'Partial'], shell=False)

def checkInput():
    # print(batdir)
    subprocess.call([batdir + '\\checkInput.bat', 'Partial'], shell=False)

def openBooksAll():
    subprocess.call([batdir + '\\openBooks.bat', 'All'], shell=False)

def openDeposit():
    subprocess.call([batdir + '\\openDeposit.bat'], shell=False)

def sendAliyos():
    subprocess.call([batdir + '\\aliyos.bat'], shell=False)

def sendMishberech():
    subprocess.call([batdir + '\\sendMish.bat'], shell=False)

def openMishberech():
    subprocess.call([batdir + '\\mishFile.bat'], shell=False)

def getFilenames():
    f = []
    mypath = basedir

    for (dirpath, dirnames, filenames) in walk(mypath):
        f.extend(filenames)
    for x in f:
        if (x == 'yahrzeits.xlsx'):
            # print("found it: " + x)
            fname = mypath + r'\\shulCloud\\yahrzeits.xlsx'
            checkOldFile(fname)

def getFileAge(fileName):
    # print(fileName)
    try:
        fileDate = time.ctime(os.path.getmtime(fileName))
        currDateDt = datetime.now()
        fileDateDt = datetime.strptime(fileDate, '%a %b %d %H:%M:%S %Y')
        return (currDateDt - fileDateDt).days
    except:
        return 365


def noExcelFile(fname):
    f = []
    for (dirpath, dirnames, filenames) in walk(basedir):
        f.extend(filenames)
    for x in f:
        # if("trans" in x):
        #     print("fname: [" + fname + "] x:" + x + ":")
        if (x == fname + '.xlsx'):
            fname = basedir + r'\\shulCloud\\' + fname + '.xlsx'
            fAge = getFileAge(fname)
            if fAge > 25:
                os.remove(fname)
                return True

            return False

    return True

def fixFile(fname):
    f = open(downloadPath + '\\' + fname + '.csv')
    reader = csv.reader(f)

    wb = Workbook()
    dest_filename = basedir + r'\\shulCloud\\' + fname + '.xlsx'

    ws = wb.worksheets[0]
    ws.title = fname

    wsRow = 2
    for row_index, row in enumerate(reader):
        wsCol = 1
        for column_index, cell in enumerate(row):
            ws.cell(row=wsRow,column=wsCol).value = cell
            wsCol += 1
        wsRow += 1

    wb.save(filename = dest_filename)
    setFileSwitches()
    buildScreen()

####################################################################################
def doYahr01():
    # getFilenames()
    runYahr()


def buildScreen():
    global labelMessage

    if noYahrzeitFIXSw:
        buttonD1 = Button(frame, text='Build Monthly Yahrzeit List', bg='yellow', fg='red')
        buttonD1R = Button(frame, text='FIX', bg='yellow', fg='red', command=partial(fixFile, 'yahrzeits'))
        buttonD1R.bind("<Enter>", writeMessageFix)
        buttonD1R.bind("<Leave>", eraseMessage)
        buttonD1R.pack(side=RIGHT)
        buttonD1R.place(x=285, y=50, bordermode=OUTSIDE, height=30, width=30)
        buttonD2 = Button(frame, text='Yahrzeit Names for Bulletin', bg='tan', fg='red')
    else:
        buttonD1 = Button(frame, text='Build Monthly Yahrzeit List', bg='yellow',  fg='black', command=doYahr01)
        buttonD1R = Button(frame, text='OK', bg='yellow', fg='black')
        buttonD1R.pack(side=RIGHT)
        buttonD1R.place(x=285, y=50, bordermode=OUTSIDE, height=30, width=30)
        buttonD2 = Button(frame, text='Yahrzeit Names for Bulletin', bg='tan',  fg='black', command=getNames)

    buttonD1.bind("<Enter>", writeMessage01)
    buttonD1.bind("<Leave>", eraseMessage)
    buttonD2.bind("<Enter>", writeMessage02)
    buttonD2.bind("<Leave>", eraseMessage)

    buttonD3L = Button(frame, text='Check View', bg='teal', command=openBooks)
    buttonD3L.bind("<Enter>", writeMessage03L)
    buttonD3L.bind("<Leave>", eraseMessage)

    buttonD3R = Button(frame, text='Check View (All)', bg='teal', command=openBooksAll)
    buttonD3R.bind("<Enter>", writeMessage03R)
    buttonD3R.bind("<Leave>", eraseMessage)

    buttonD3RR = Button(frame, text='Move Chk', bg='teal', command=checkMove)
    buttonD3RR.bind("<Enter>", writeMessage03RR)
    buttonD3RR.bind("<Leave>", eraseMessage)

    buttonD4 = Button(frame, text='Open Deposit View', bg='purple', command=openDeposit)
    buttonD4R = Button(frame, text='File', bg='purple', fg='black', command=partial(openFile, 'Deposits'))
    buttonD4R.pack(side=RIGHT)
    buttonD4R.place(x=285, y=210, bordermode=OUTSIDE, height=30, width=60)
    buttonD4R.bind("<Enter>", writeMessage05R)
    buttonD4R.bind("<Leave>", eraseMessage)

    buttonD5 = Button(frame, text='Enter Checks/Cash', bg='teal',  fg='black', command=checkInput)
    buttonD5.pack(side=RIGHT)
    buttonD5.place(x=75, y=170, bordermode=OUTSIDE, height=30, width=200)
    buttonD5R = Button(frame, text='File', bg='teal', fg='black', command=partial(openFile, 'DailyLog'))
    buttonD5R.bind("<Enter>", writeMessage05R)
    buttonD5R.bind("<Leave>", eraseMessage)

    buttonD5R.pack(side=RIGHT)
    buttonD5R.place(x=285, y=170, bordermode=OUTSIDE, height=30, width=60)

    if noYahrzeitFileSw:
        message = "Please DOWNLOAD and FIX your 'yahrzeits' file"
        labelB14 = Label(frame, text=message, bg="black", fg="red", font='Helvetica 8')
        labelB14.pack(side=RIGHT)
        labelB14.place(x=1, y=1, bordermode=OUTSIDE, height=40, width=270)
    else:
        if noTransactionFileSw:
            message = "Please DOWNLOAD and FIX your 'transactions' file"
            labelB14 = Label(frame, text=message, bg="black", fg="red", font='Helvetica 8')
            labelB14.pack(side=RIGHT)
            labelB14.place(x=1, y=1, bordermode=OUTSIDE, height=40, width=270)

    if noTransactionFIXSw:
        buttonD14 = Button(frame, text='Send Email for aliyos', bg='pink', fg='red')
        buttonD14R = Button(frame, text='FIX', bg='yellow', fg='red', command=partial(fixFile, 'transactions'))
        buttonD14R.pack(side=RIGHT)
        buttonD14R.place(x=285, y=270, bordermode=OUTSIDE, height=30, width=30)
    else:
        buttonD14 = Button(frame, text='Send Email for aliyos', bg='pink', fg='black', command=sendAliyos)
        buttonD14R = Button(frame, text='OK', bg='pink', fg='black')
        buttonD14R.pack(side=RIGHT)
        buttonD14R.place(x=285, y=270, bordermode=OUTSIDE, height=30, width=30)

    buttonD14.bind("<Enter>", writeMessage14R)
    buttonD14.bind("<Leave>", eraseMessage)

    buttonD15 = Button(frame, text='Send Email for mishberech', bg='orange', command=sendMishberech)
    buttonD15R = Button(frame, text='File', bg='orange', fg='black', command=openMishberech)
    buttonD15R.pack(side=RIGHT)
    buttonD15R.place(x=285, y=310, bordermode=OUTSIDE, height=30, width=60)
    buttonD15.bind("<Enter>", writeMessage14R)
    buttonD15.bind("<Leave>", eraseMessage)

    buttonD1.pack(side=RIGHT)
    buttonD2.pack(side=RIGHT)
    buttonD3L.pack(side=RIGHT)
    buttonD3R.pack(side=RIGHT)
    buttonD3RR.pack(side=RIGHT)
    buttonD4.pack(side=RIGHT)

    buttonD14.pack(side=RIGHT)
    buttonD15.pack(side=RIGHT)

    buttonD1.place(x=75, y=50, bordermode=OUTSIDE, height=30, width=200)
    buttonD2.place(x=75, y=90, bordermode=OUTSIDE, height=30, width=200)
    buttonD3L.place(x=75, y=130, bordermode=OUTSIDE, height=30, width=100)
    buttonD3R.place(x=175, y=130, bordermode=OUTSIDE, height=30, width=100)
    buttonD3RR.place(x=285, y=130, bordermode=OUTSIDE, height=30, width=60)
    buttonD4.place(x=75, y=210, bordermode=OUTSIDE, height=30, width=200)

    buttonD14.place(x=75, y=270, bordermode=OUTSIDE, height=30, width=200)
    buttonD15.place(x=75, y=310, bordermode=OUTSIDE, height=30, width=200)

    labelMessage = Label(frame, text="...message...", bg="black", fg="yellow", font='Helvetica 8')
    labelMessage.pack(side=RIGHT)
    labelMessage.place(x=1, y=350, bordermode=OUTSIDE, height=30, width=270)
    labelMessage.bind("<Enter>", writeMessage)
    labelMessage.bind("<Leave>", eraseMessage)

def writeMessage(event):
    labelMessage.config(text="Stam... message")

def writeMessageFix(event):
    labelMessage.config(text="First download a new file from shulCloud, then press")

def writeMessage01(event):
    labelMessage.config(text="Use about 1 week before the Hebrew month")

def writeMessage02(event):
    labelMessage.config(text="Gets all yarhzeit names for the next week")

def writeMessage03L(event):
    labelMessage.config(text="Displays recent checks by deposit, person, and date")

def writeMessage03R(event):
    labelMessage.config(text="Displays ALL checks by deposit, person, and date")

def writeMessage03RR(event):
    labelMessage.config(text="Exactly 1 file scanned - this moves the file to 'Checks'")

def writeMessage05R(event):
    labelMessage.config(text="open the EXCEL file")

def writeMessage14R(event):
    labelMessage.config(text="SENDS Email to members!")

def eraseMessage(event):
    labelMessage.config(text="")

def setFileSwitches():
    global noYahrzeitFileSw
    global noYahrzeitFIXSw
    global noTransactionFileSw
    global noTransactionFIXSw

    noYahrzeitFileSw = False
    noYahrzeitFIXSw = False
    noTransactionFileSw = False
    noTransactionFIXSw = False

    if noExcelFile('yahrzeits'):
        noYahrzeitFIXSw = True
        downloadedCSV = downloadPath + "\\yahrzeits.csv"
        fileAge = getFileAge(downloadedCSV)
        if fileAge > 25:
            noYahrzeitFileSw = True

    if noExcelFile('transactions'):
        noTransactionFIXSw = True
        downloadedCSV = downloadPath + "\\transactions.csv"
        fileAge = getFileAge(downloadedCSV)
        if fileAge > 25:
            noTransactionFileSw = True

    # downloadedCSV = downloadPath + "transactions.csv"
    # fileAge = getFileAge(downloadedCSV)
    # if fileAge > 1:
    #     noTransactionFileSw = True

frame = Frame(root, width=400, height=400)

setFileSwitches()
buildScreen()
frame.configure(background='black')
frame.pack()
root.mainloop()
