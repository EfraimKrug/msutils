import csv
from openpyxl import Workbook
#from openpyxl.cell import get_column_letter

downloadPath = r'C:\\Users\\KTM\\Downloads\\'

# def get_column_letter(num):
#     alpha = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L', 'M', 'N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
#     if(num < 26):
#         return alpha[num-1]
#     return alpha[num - 26] + alpha[num % 26]

f = open(downloadPath + 'yahrzeits.csv')

#csv.register_dialect('colons', delimiter=':')

#reader = csv.reader(f, dialect='colons')
reader = csv.reader(f)

wb = Workbook()
bd = r'C:\\Users\\KTM\\python\\msutils'
dest_filename = bd + r'\\shulCloud\\yahrzeits.xlsx'
# dest_filename = 'yahrzeits.xlsx'

ws = wb.worksheets[0]
ws.title = "yahrzeits"

wsRow = 2
for row_index, row in enumerate(reader):
    wsCol = 1
    for column_index, cell in enumerate(row):
        # column_letter = get_column_letter((column_index + 1))

        # print(cell)
        ws.cell(row=wsRow,column=wsCol).value = cell
        wsCol += 1
    wsRow += 1
        #print(column_letter)
        #ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

wb.save(filename = dest_filename)
