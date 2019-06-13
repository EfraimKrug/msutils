#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
#########################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import csv
import shutil
import requests

from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from AlefBet import *
from periodProcess import *
from mishDisplay01 import *

fromaddr = 'KadimahTorasMoshe@gmail.com'
toaddrs  = 'EfraimMKrug@gmail.com'
server = ''
wbook = ''
####################################################
#   This process is the update for the MishBerech
#   excel spread sheet
####################################################

last = 0
accountArray = []
accountArray.append(dict())

def downloadXLSX():
    url = "https://images.shulcloud.com/616/uploads/mishberech/MishBerech.xlsx"
    response = requests.get(url, stream=True)
    with open('MishBerech.xlsx', 'wb') as out_file:
        shutil.copyfileobj(response.raw, out_file)
    del response

def openTrx():
    global wbook
    wbook = load_workbook('MishBerech.xlsx')
    return wbook

def fixName(name):
    if ord(name[0]) > 500:
        fixedName = name[::-1]
    else:
        fixedName = name
    return fixedName


def getTrx(sheet):
    global accountArray # array of dict each key is a shul member's name
                        # accountArray[0]['name'] = [] // array of cholim => [name, date, notes, email]
                        # i.e. accountArray[0]['name'][0][0] is the name of a chole from date accountArray[0]['name'][0][1]
    idx = 0
    hold = dict()
    track = []
    for r in range(4, sheet.max_row + 1):
        if(len(str(sheet.cell(row=r, column=3).value).strip()) < 5):
            continue

        hold = dict()
        if(not str(sheet.cell(row=r, column=2).value).strip() in track):
            accountArray.append(dict())
            accountArray[len(accountArray)-1][sheet.cell(row=r, column=2).value] = []
            track.append(str(sheet.cell(row=r, column=2).value).strip())

        for dct in accountArray:
            if str(sheet.cell(row=r, column=2).value) in dct:
                hold = dct

        hold[sheet.cell(row=r, column=2).value].append([])  #name
        hold[sheet.cell(row=r, column=2).value][len(hold[sheet.cell(row=r, column=2).value])-1].append(fixName(sheet.cell(row=r, column=3).value))
        hold[sheet.cell(row=r, column=2).value][len(hold[sheet.cell(row=r, column=2).value])-1].append(sheet.cell(row=r, column=4).value)  #date
        hold[sheet.cell(row=r, column=2).value][len(hold[sheet.cell(row=r, column=2).value])-1].append(sheet.cell(row=r, column=5).value)  #comment
        hold[sheet.cell(row=r, column=2).value][len(hold[sheet.cell(row=r, column=2).value])-1].append(sheet.cell(row=r, column=6).value)  #email

def saveData():
    #print("Saving The Data... as it were")
    global accountArray # array of dict each key is a shul member's name
                        # accountArray[0]['name'] = [] // array of cholim => [name, date, notes, email]
                        # i.e. accountArray[0]['name'][0][0] is the name of a chole from date accountArray[0]['name'][0][1]
    global newSheet
    newSheet = wbook['Rebuild']

    newSheet.cell(row=1, column=3).value = "Misheberach List"

    newSheet.cell(row=3, column=2).value = "Person Requesting"
    newSheet.cell(row=3, column=3).value = "Name and Mother's Name"
    newSheet.cell(row=3, column=4).value = "Date Added"
    newSheet.cell(row=3, column=5).value = "Comments"
    newSheet.cell(row=3, column=6).value = "Email Address"

    row_num = 4
    for dct in accountArray:
        for d in dct:
            #print ("===>" + d)
            newSheet.cell(row=row_num, column=2).value = d
            for ent in dct[d]:
                #print("============> " + ent[0])
                newSheet.cell(row=row_num, column=3).value = ent[0]
                newSheet.cell(row=row_num, column=4).value = ent[1]
                newSheet.cell(row=row_num, column=5).value = ent[2]
                newSheet.cell(row=row_num, column=6).value = ent[3]
                row_num += 1

    wbook.save('new.xlsx')

def printDisplayAccounts():
    for d in accountArray:
        print("#"*50)
        for x in d:
            print("Member: " + x)
            for a in d[x]:
                print("\tName: " + str(a[0]))
                print("\tDate: " + str(a[1]))
                print("\tNotes: " + str(a[2]))
                print("\tEmail: " + str(a[3]))
#######################################################\

def runProcess():
    global server
    global wbook

    downloadXLSX()
    wbook = openTrx()
    keys = []
    getTrx(wbook['Updates'])
    newSheet = wbook.create_sheet(title = 'Rebuild')
    newSheet = wbook.active

    printDisplayAccounts()

def main():
    runProcess()
    root = tk.Tk()
    app = mishDisplay01(root, accountArray, saveData)
    root.mainloop()

if __name__ == '__main__':
    main()
