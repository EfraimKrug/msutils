#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
#########################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import csv
import shutil
import requests

from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from AlefBet import *
from periodProcess import *
import string

fromaddr = 'KadimahTorasMoshe@gmail.com'
toaddrs  = 'EfraimMKrug@gmail.com'
server = ''
fileName = "MishBerech.xlsx"

reload(sys)
sys.setdefaultencoding('utf-8')
####################################################
#username = 'KadimahTorasMoshe@gmail.com'
#password = 'August7Brachas'
####################################################

last = 0
accountArray = []
accountArray.append(dict())

def downloadXLSX():
    url = "https://images.shulcloud.com/616/uploads/mishberech/MishBerech.xlsx"
    response = requests.get(url, stream=True)
    with open(fileName, 'wb') as out_file:
        shutil.copyfileobj(response.raw, out_file)
    del response

def openTrx():
    wb = load_workbook(fileName)
    return wb

def fixName(name):
    if ord(name[0]) > 500:
        fixedName = name[::-1]
    else:
        fixedName = name
    return fixedName


def getTrx(sheet):
    global last
    for r in range(4, sheet.max_row + 1):
        if(len((str(sheet.cell(row=r, column=2).value)).strip()) < 1 or
            sheet.cell(row=r, column=6).value is None):
            continue

        if sheet.cell(row=r, column=6).value.find("@") > -1:
            accountArray.append(dict())
            last += 1

        accountArray[last][r] = []
        accountArray[last][r].append(sheet.cell(row=r, column=2).value)  #name
        accountArray[last][r].append(fixName(sheet.cell(row=r, column=3).value))
        accountArray[last][r].append(sheet.cell(row=r, column=4).value)  #date
        accountArray[last][r].append(sheet.cell(row=r, column=5).value)  #comment
        accountArray[last][r].append(sheet.cell(row=r, column=6).value)  #email

def checkDays():
    today_dt = datetime.now()

    for accounts in accountArray:
        for id in accounts:
            line_dt = datetime.strptime(str(accounts[id][2])[0:10], "%Y-%m-%d")
            dayCount = (today_dt - line_dt).days
            if dayCount < 45:
                accounts[id][4] = ""

def writeEmail(account):
    s = ""
    s = "Dear " + account[0] + ",\n\n"
    s += "We have been saying a mish'berech for " + getUsableWord(account[1])
    s += "\nsince " + str(account[2])[0:10] + ".\n\n"
    s += "Please email us if you would like us to continue mentioning this name in the shul"
    s += "\nmish'berech, otherwise we will be removing the name after this coming Shabbos."
    s += "\n\nThanks so much!\n\nBest and Good Shabbos! "
    return s

def fire(fromaddr, toaddrs, msg):
    global server
    p = set(string.printable)

    try:
        server.sendmail(fromaddr, filter(lambda x: x in p, toaddrs), msg)
        print("Successful Send: " + filter(lambda x: x in p, toaddrs))
        return True
    except Exception as e:
        print("#OUCH" + ":[" + filter(lambda x: x in p, toaddrs) + "]: " + str(e))

    return False

def updateExcel(sheet, account):
        today_dt = datetime.now()
        for r in range(4, sheet.max_row + 1):
            if account[0] == sheet.cell(row=r, column=2).value and account[4] == sheet.cell(row=r, column=6).value:
                sheet.cell(row=r, column=5).value = today_dt
                return



def sendItAll(sheet):
    global server
    global fromaddr
    global toaddrs

    usedAddressList = []
    for accounts in accountArray:
        for a in accounts:
            msgTxt = writeEmail(accounts[a])
            toaddrs = accounts[a][4]
            #toaddrs = "efraimmkrug@gmail.com"
            if toaddrs.find('@') < 0:
                continue
            x = accounts[a][4]
            msg = "\r\n".join([
              "From: " + fromaddr,
              "To: " + x.encode('ascii', 'ignore'),
              "Subject: MishBerech list",
              "",
              msgTxt
              ])

            #print("FROM: " + fromaddr)
            #print("TO: " + x.encode('ascii', 'ignore'))
            if fire(fromaddr, toaddrs, msg):
                updateExcel(sheet, accounts[a])


#######################################################\
def runProcess():
    global server
    downloadXLSX()
    trx = openTrx()
    getTrx(trx[trx.sheetnames[0]])
    checkDays()
    server = smtplib.SMTP(smtpvar)
    server.ehlo()
    server.starttls()
    server.login(username,password)
    sendItAll(trx[trx.sheetnames[0]])
    server.quit()
    trx.save(fileName)

runMonthly(runProcess)
