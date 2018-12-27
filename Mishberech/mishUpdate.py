#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from profile import *

fromaddr = 'KadimahTorasMoshe@gmail.com'
toaddrs  = 'EfraimMKrug@gmail.com'

####################################################
username = 'KadimahTorasMoshe@gmail.com'
password = 'August7Brachas'
####################################################

last = 0
accountArray = []
accountArray.append(dict())

def openTrx():
    wb = load_workbook('MishBerech.xlsx')
    return wb

def fixName(name):
    if ord(name[0]) > 500:
        fixedName = name[::-1]
    else:
        fixedName = name
    return fixedName


def getTrx(sheet):
    global last
    for r in range(7, sheet.max_row + 1):
        if(type(sheet.cell(row=r, column=6).value) != unicode):
            continue
        if sheet.cell(row=r, column=6).value.find("@") > -1:
            accountArray.append(dict())
            last += 1

        accountArray[last][r] = []
        accountArray[last][r].append(sheet.cell(row=r, column=2).value)  #name
        accountArray[last][r].append(fixName(sheet.cell(row=r, column=3).value))
        accountArray[last][r].append(sheet.cell(row=r, column=4).value)  #date
        accountArray[last][r].append(sheet.cell(row=r, column=5).value)  #comment
        accountArray[last][r].append(sheet.cell(row=r, column=6).value)  #email

def checkDays():
    today_dt = datetime.now()

    for accounts in accountArray:
        for id in accounts:
            line_dt = datetime.strptime(str(accounts[id][2])[0:10], "%Y-%m-%d")
            dayCount = (today_dt - line_dt).days
            if dayCount < 45:
                accounts[id][4] = ""

def writeEmail(account):
    s = ""
    s = "Dear " + account[0] + ",\n\n"
    s += "We have been saying a mish'berech for " + account[1]
    s += "\nsince " + str(account[2])[0:10] + ".\n\n"
    s += "Please email us if you would like us to continue mentioning this name in the shul"
    s += "\nmish'berech, otherwise we will be removing the name after this coming Shabbos."
    s += "\n\nThanks so much!\n\nBest and Good Shabbos! "
    return s

def fire(fromaddr, toaddrs, msg):
    #server.sendmail(fromaddr, toaddrs, msg)
    print("#"*45)
    print(msg)
    print("#"*45)


def sendItAll():
    global server
    global fromaddr
    global toaddrs

    usedAddressList = []
    for accounts in accountArray:
        for a in accounts:
            msgTxt = writeEmail(accounts[a])
            toaddrs = accounts[a][4]
            if toaddrs.find('@') < 0:
                continue
            msg = "\r\n".join([
              "From: " + fromaddr,
              "To: " + accounts[a][4],
              "Subject: MishBerech list",
              "",
              msgTxt
              ])
            fire(fromaddr, toaddrs, msg)


#######################################################\
trx = openTrx()
getTrx(trx[trx.sheetnames[0]])
checkDays()
server = smtplib.SMTP(smtpvar)
server.ehlo()
server.starttls()
server.login(username,password)
sendItAll()
#server.quit()
