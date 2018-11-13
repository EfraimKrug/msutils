#Recon001.py
import sys
#import webbrowser
import csv
from openpyxl import load_workbook
import winreg
from profile import *

PP_TRANS = dict()
SC_TRANS = dict()

CHECKS = dict()
#CHROME_PATH = 'C:/Users/KTM/AppData/Local/Google/Chrome/Application/chrome.exe %s'
URLS = []

wDir = basedir + "\\Reconcile"

def openPP():
  wb = load_workbook(wDir + "\\PPTrans.xlsx")
  return wb

def openSC():
  wb = load_workbook(wDir + "\\SCTrans.xlsx")
  return wb

def getPayPal(wbook):
    _num = ""

    sheet = wbook[wbook.sheetnames[0]]
    for r in range(2, sheet.max_row+1):
        if(sheet.cell(row=r,column=11).value != "office@ktmshul.org"):
            _num = sheet.cell(row=r,column=13).value
            PP_TRANS[_num] = dict()
            PP_TRANS[_num]['date'] = str(sheet.cell(row=r,column=1).value)
            PP_TRANS[_num]['time'] = sheet.cell(row=r,column=2).value
            PP_TRANS[_num]['name'] = sheet.cell(row=r,column=4).value
            PP_TRANS[_num]['gross'] = str(sheet.cell(row=r,column=8).value)
            PP_TRANS[_num]['fee'] = sheet.cell(row=r,column=9).value
            PP_TRANS[_num]['email'] = sheet.cell(row=r,column=11).value

def getSC(wbook):
    _num = ""

    sheet = wbook[wbook.sheetnames[0]]
    for r in range(2, sheet.max_row+1):
        if(len(str(sheet.cell(row=r,column=7).value)) > 2):
            _num = str(sheet.cell(row=r,column=7).value)
            SC_TRANS[_num] = dict()
            SC_TRANS[_num]['date'] = sheet.cell(row=r,column=1).value
            SC_TRANS[_num]['type'] = sheet.cell(row=r,column=6).value
            SC_TRANS[_num]['name'] = sheet.cell(row=r,column=3).value
            SC_TRANS[_num]['gross'] = sheet.cell(row=r,column=10).value
            SC_TRANS[_num]['id'] = sheet.cell(row=r,column=14).value
            #SC_TRANS[_num]['email'] = sheet.cell(row=r,column=11).value

def printPP():
    for t in PP_TRANS:
        print(t + "::" + PP_TRANS[t]['email'])

def printSC():
    for t in SC_TRANS:
        print(t + "::" + SC_TRANS[t]['name'])

def getPPnoSC():
    ###################################################################
    # This function will report on any transaction in PayPal that
    # has NOT been logged on ShulCloud
    ###################################################################

    yesCount = 0
    noCount = 0

    for _num in PP_TRANS:
        if _num in SC_TRANS:
            #print(_num + "... reconciled")
            yesCount += 1
        else:
            #print(_num + " no shul cloud entry!")
            print("[" + str(_num) + "]::" + str(PP_TRANS[_num]['date']) + "::" + PP_TRANS[_num]['name'] + "::" + PP_TRANS[_num]['email'] + "::" + PP_TRANS[_num]['gross'])
            noCount += 1

    print (str(noCount) + " Failures to reconcile")
    print (str(yesCount) + " Reconciliations")

def getSCnoPP():
    ###################################################################
    # This function will report on any transaction in ShulCloud that
    # has NOT been logged on PayPal
    ###################################################################

    yesCount = 0
    noCount = 0

    for _num in SC_TRANS:
        if SC_TRANS[_num]['type'].lower().find("paypal") > -1:
            if _num in PP_TRANS:
                #print(_num + "... reconciled")
                yesCount += 1
            else:
                #print(_num + " no shul cloud entry!")
                print("[" + str(_num) + "]::" + str(SC_TRANS[_num]['date']) + "::" + SC_TRANS[_num]['name'] + "::" + str(SC_TRANS[_num]['gross']))
                noCount += 1

    print (str(noCount) + " Failures to reconcile")
    print (str(yesCount) + " Reconciliations")

getPayPal(openPP())
getSC(openSC())
print ("****************** Pay Pal - with no transaction in Shul Cloud *************************")
getPPnoSC()
print ("****************** Shul Cloud - with no transaction in Pay Pal *************************")
getSCnoPP()
#printSC()
