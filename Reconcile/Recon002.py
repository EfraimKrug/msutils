#Recon002.py
import sys
import csv
from openpyxl import load_workbook
import winreg
from profile import *

PP_TRANS = dict()
SC_TRANS = dict()
EMAIL = dict()

CHECKS = dict()
#CHROME_PATH = 'C:/Users/KTM/AppData/Local/Google/Chrome/Application/chrome.exe %s'
URLS = []

wDir = basedir + "\\Reconcile"


def openPP():
  wb = load_workbook("..\\shulCloud\\PeoplePP.xlsx")
  return wb

def openSC():
  wb = load_workbook("..\\shulCloud\\PeopleSC.xlsx")
  return wb

def getPayPal(wbook):
    _email = ""

    sheet = wbook[wbook.sheetnames[0]]
    for r in range(2, sheet.max_row+1):
        if(sheet.cell(row=r,column=11).value != "office@ktmshul.org"):
            _email = sheet.cell(row=r,column=11).value
            PP_TRANS[_email] = dict()
            PP_TRANS[_email]['name'] = sheet.cell(row=r,column=4).value
            EMAIL[_email] = str(PP_TRANS[_email]['name'])
            #PP_TRANS[_num]['email'] = sheet.cell(row=r,column=11).value

def getSC(wbook):
    _email = ""

    sheet = wbook[wbook.sheetnames[0]]
    for r in range(2, sheet.max_row+1):
        if(len(str(sheet.cell(row=r,column=33).value)) > 2):
            _email = str(sheet.cell(row=r,column=33).value)
            SC_TRANS[_email] = dict()
            SC_TRANS[_email]['id'] = sheet.cell(row=r,column=1).value
            SC_TRANS[_email]['lname'] = sheet.cell(row=r,column=3).value
            SC_TRANS[_email]['fname'] = sheet.cell(row=r,column=4).value
            EMAIL[_email] = str(SC_TRANS[_email]['lname']) + ", " + str(SC_TRANS[_email]['fname'])
            #SC_TRANS[_num]['email'] = sheet.cell(row=r,column=33).value

def printPP():
    for t in PP_TRANS:
        print(t + "::" + PP_TRANS[t]['email'])

def printSC():
    for t in SC_TRANS:
        print(t + "::" + SC_TRANS[t]['name'])

def reformatName(name):
    n = name.strip()
    chr = ' '

    if n.find(",") > -1:
        chr = ','

    n1 = n[n.find(chr)+1:]
    n2 = n[0:n.find(chr)]
    if chr == ' ':
        n3 = n2
        n2 = n1
        n1 = n3

    #print ("====>>> " + n + ": {" + n1 + "}[" + n2 + "]")
    n = n2.title() + ", " + n1.title()
    return n.replace(",  ", ", ")

def getPPnoSC():
    print ("*********************************************************")
    print ("** List of email addresses from our PayPal transactions *")
    print ("** that we do not have in ShulCloud                     *")
    print ("*********************************************************")
    yesCount = 0
    noCount = 0

    for _email in PP_TRANS:
        if _email in SC_TRANS:
            #print(_num + "... reconciled")
            yesCount += 1
        else:
            print(str(PP_TRANS[_email]['name']) + " <" + _email +">")
            noCount += 1

    #print (str(noCount) + " Failures to reconcile")
    #print (str(yesCount) + " Reconciliations")

def printEMAIL(fout):
    temp = dict()
    temp2 = dict()
    print ("*********************************************************")
    print ("** List of email addresses from our PayPal transactions *")
    print ("** and from our ShulCloud transactions                  *")
    print ("** Sorted according to email                            *")
    print ("*********************************************************")
    for _email in EMAIL:
        name = reformatName(EMAIL[_email])
        temp[_email.lower()] = name
        temp2[name] = _email

    for _email in sorted(temp.iterkeys()):
        print "* " + _email + " <" + temp[_email] + ">"
        fout.write(_email + "," + temp[_email] + "\n")

    print ("*********************************************************")
    print ("** Sorted according to names                            *")
    print ("*********************************************************")

    for name in sorted(temp2.iterkeys()):
        print "* " + name + " <" + temp2[name] + ">"
        fout.write(name + "," + temp2[name] + "\n")

    print ("*********************************************************")

f = open(wDir + "\\emailList.csv", "w")
getPayPal(openPP())
getSC(openSC())

getPPnoSC()
printEMAIL(f)
f.close()
