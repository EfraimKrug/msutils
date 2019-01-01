# get names for the weekly bulletin
#getNames.py
#
import sys
import os
#########################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import webbrowser
import winreg

from openpyxl import load_workbook

from pyluach import dates, hebrewcal
from Profile import *

next_next_next_shabbos = ''
next_next_shabbos = ''
next_shabbos = ''
today = ''
#shabbos_greg = ''

LINE_WIDTH = 75

def setDates():
        global next_next_next_shabbos
        global next_next_shabbos
        global next_shabbos
        global today

        today = dates.HebrewDate.today()
        remaining = 7 - today.weekday()
        next_shabbos = today + remaining
        #shabbos_greg = next_shabbos.to_greg()
        next_next_shabbos = today + remaining + 7
        next_next_next_shabbos = today + remaining + 14

def getWorkBook():
     wb = load_workbook(basedir + '\\shulCloud\\yahrzeits.xlsx')
     return wb

def getMonthNum(dt):
    if(dt.lower().find('nis') > -1): return 1
    if(dt.lower().find('iyar') > -1): return 2
    if(dt.lower().find('sivan') > -1): return 3
    if(dt.lower().find('tam') > -1): return 4
    if(dt.lower().find('av') > -1): return 5
    if(dt.lower().find('elul') > -1): return 6
    if(dt.lower().find('tishr') > -1): return 7
    if(dt.lower().find('heshvan') > -1): return 8
    if(dt.lower().find('islev') > -1): return 9
    if(dt.lower().find('teve') > -1): return 10
    if(dt.lower().find('shva') > -1): return 11
    if(dt.lower().find('adar') > -1):
        if(dt.lower().find('II') > -1): return 13
        return 12
    return 0

def getDay(dt):
    #print("[" + str(dt) + "]")

    try:
        d = int(dt.strip()[0:dt.find(' ')])
    except ValueError:
        d = 1

    if isinstance(d, int):
        return d
    return 1

def getHebDate(m, d):
    global today
    if(m == 0):
        return today

    return dates.HebrewDate(today.year, m, d)

def compareDates(dt1, dt2):
    if dt1 == today:
        return -2
    if(dt1 > dt2):
        return -1
    if(dt2 > dt1):
        return 1
    return 0

def getNames(sheet):
    names = []
    i = 1
    for r in range(2, sheet.max_row):
        dt = str(sheet.cell(row=r,column=8).value)        #print("current: " + str(dt) + "::" + str(getHebDate(getMonthNum(dt), getDay(dt))) )
        if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_shabbos) in [-1,0]):
            if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_next_shabbos) in [0, 1]):                #print (str(i) + ") " + str(sheet.cell(row=r,column=2).value + "::" + dt))
                names.append(str(sheet.cell(row=r,column=2).value))
                i+=1
    return names

def getNextNames(sheet):
    names = []
    i = 1
    for r in range(2, sheet.max_row):
        dt = str(sheet.cell(row=r,column=8).value)        #print("current: " + str(dt) + "::" + str(getHebDate(getMonthNum(dt), getDay(dt))) )
        if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_next_shabbos) in [-1,0]):
            if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_next_next_shabbos) in [0, 1]):                #print (str(i) + ") " + str(sheet.cell(row=r,column=2).value + "::" + dt))
                names.append(str(sheet.cell(row=r,column=2).value))
                i+=1
    return names

def insertBreak(breakPos, txt):
    j = 0
    s = ""
    while j < breakPos:
        s += txt[j]
        j += 1
    s += "\n"
    j += 1
    while j < len(txt) + 1:
        s += txt[j-1]
        j+=1
    return s

def fixLineWidths(txt):
    i = LINE_WIDTH
    lastBreak = 0
    while i < len(txt):
        while txt[i] != ";" and i > lastBreak:
            i -= 1
        txt = insertBreak(i+1, txt)
        lastBreak = i + 1
        i += LINE_WIDTH
    return txt

def checkString(txt):
    i = 0
    j = 0
    while i < len(txt):
        if txt[i] == ";":
            j = 0
        if j > (LINE_WIDTH - 2):
            return 0
        j += 1
        i += 1
    return 1

def openFile(fname):
    f = open(basedir + "\\newFiles\\" + fname + ".txt", "w")
    return f

def printToFile(names, f):
    s = "  "
    for n in names:
        if len(s) > 2:
            s += ";  "
        s += n
    if checkString(s) < 1:
        print("Sorry - at least one name is longer than the maximum line width")
        return

    t = fixLineWidths(s)
    f.write(t)
    f.close()

setDates()
wbook = getWorkBook()
sheet = wbook[wbook.sheetnames[0]]
n = getNames(sheet)
n2 = getNextNames(sheet)

printToFile(n, openFile("NamesWeek01"))
printToFile(n2, openFile("NamesWeek02"))

#today_dt = str(next_shabbos.to_greg().month) + "-" + str(next_shabbos.to_greg().day) + "-" + str(next_shabbos.to_greg().year)
