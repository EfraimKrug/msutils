# get names for the weekly bulletin
#getNames.py
#
import sys
#import subprocess
import webbrowser
import winreg

from openpyxl import load_workbook

from pyluach import dates, hebrewcal

latestShmaURL_1 = "https://www.myzmanim.com/day.aspx?vars=72080767/"    # 11-24-2018
latestShmaURL_2 = "////////////////6ccdf7"

next_next_shabbos = ''
next_shabbos = ''
today = ''
shabbos_greg = ''

LINE_WIDTH = 75

WORDEXE = r'C:\Program Files\Microsoft Office\root\Office\WINWORD.EXE'
CHROME = 'C:/Users/KTM/AppData/Local/Google/Chrome/Application/chrome.exe %s'

def getChrome():
    global CHROME_PATH
    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe")

    num_values = winreg.QueryInfoKey(handle)[1]
    for i in range(num_values):
        for x in winreg.EnumValue(handle, i):
            if(str(x).find("CHROME") > -1):
                CHROME_PATH = x + " %s"

def getWord():
    global WORDEXE
    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\winword.exe")

    num_values = winreg.QueryInfoKey(handle)[1]
    for i in range(num_values):
        for x in winreg.EnumValue(handle, i):
            if(str(x).find("WINWORD") > -1):
                WORDEXE = x

def setDates():
        global next_next_shabbos
        global next_shabbos
        global today

        today = dates.HebrewDate.today()
        remaining = 7 - today.weekday()
        next_shabbos = today + remaining
        shabbos_greg = next_shabbos.to_greg()
        next_next_shabbos = today + remaining + 7

def getWorkBook():
     wb = load_workbook('./yahrzeits.xlsx')
     return wb

def getMonthNum(dt):
    if(dt.lower().find('nissan') > -1): return 1
    if(dt.lower().find('iyar') > -1): return 2
    if(dt.lower().find('sivan') > -1): return 3
    if(dt.lower().find('tamuz') > -1): return 4
    if(dt.lower().find('av') > -1): return 5
    if(dt.lower().find('elul') > -1): return 6
    if(dt.lower().find('tishrei') > -1): return 7
    if(dt.lower().find('cheshvan') > -1): return 8
    if(dt.lower().find('kislev') > -1): return 9
    if(dt.lower().find('teves') > -1): return 10
    if(dt.lower().find('shvat') > -1): return 11
    if(dt.lower().find('adar') > -1):
        if(dt.lower().find('II') > -1): return 13
        return 12
    return 0

def getDay(dt):
    #print("[" + str(dt) + "]")

    try:
        d = int(dt.strip()[0:dt.find(' ')])
    except ValueError:
        d = 1

    if isinstance(d, int):
        return d
    return 1

def getHebDate(m, d):
    global today
    #print (str(today.year) + "::" + str(m) + "::" + str(d))
    if(m == 0):
        return today

    return dates.HebrewDate(today.year, m, d)
    #return dates.HebrewDate(today.year, m, d)

def compareDates(dt1, dt2):
    if dt1 == today:
        return -2
    if(dt1 > dt2):
        return -1
    if(dt2 > dt1):
        return 1
    return 0

def getNames(sheet):
    names = []
    i = 1
    for r in range(2, sheet.max_row):
        dt = str(sheet.cell(row=r,column=8).value)
        if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_shabbos) in [-1,0]):
            if(compareDates(getHebDate(getMonthNum(dt), getDay(dt)), next_next_shabbos) in [0, 1]):
                #print (str(i) + ") " + str(sheet.cell(row=r,column=2).value + "::" + dt))
                names.append(str(sheet.cell(row=r,column=2).value))
                i+=1
    return names

def insertBreak(breakPos, txt):
    j = 0
    s = ""
    while j < breakPos:
        s += txt[j]
        j += 1
    s += "\n"
    j += 1
    while j < len(txt) + 1:
        s += txt[j-1]
        j+=1
    return s

def fixLineWidths(txt):
    i = LINE_WIDTH
    lastBreak = 0
    while i < len(txt):
        while txt[i] != ";" and i > lastBreak:
            i -= 1
            #print(txt[i])
        #print(i+1)
        txt = insertBreak(i+1, txt)
        lastBreak = i + 1
        i += LINE_WIDTH
        #print(txt[i] + ":" + txt[i+1])
    return txt

def checkString(txt):
    i = 0
    j = 0
    while i < len(txt):
        if txt[i] == ";":
            j = 0
        if j > (LINE_WIDTH - 2):
            return 0
        j += 1
        i += 1
    return 1

def printToFile(names):
    s = ""
    for n in names:
        s += n + ";  "
    if checkString(s) < 1:
        print("Sorry - at least one name is longer than the maximum line width")
        return

    t = fixLineWidths(s)
    #print(t)
    f = open("temp.txt", "w")
    f.write(t)
    f.close()

setDates()
wbook = getWorkBook()
sheet = wbook[wbook.sheetnames[0]]
getWord()
n = getNames(sheet)
printToFile(n)
today_dt = str(next_shabbos.to_greg().month) + "-" + str(next_shabbos.to_greg().day) + "-" + str(next_shabbos.to_greg().year)
url = latestShmaURL_1 + today_dt + latestShmaURL_2


#subprocess.call([CHROME, url])
webbrowser.get(CHROME).open(url)
