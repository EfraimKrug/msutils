#yahr.py
import sys
import os
#########################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import csv
from Profile import *

from pyluach import dates, hebrewcal

from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side
adarII = False

def getNextMonth():
    month = dates.HebrewDate.today().month + 1
    if hebrewcal.Year(dates.HebrewDate.today().year).leap:
        if month > 13:
            month = month - 13
    else:
        if month > 12:
            month = month - 12
    return month

def getYear():
    return dates.HebrewDate.today().year

def getNextMonthName():
    #return "Kislev"
    return hebrewcal.Month(getYear(), getNextMonth()).name

def openwb():
  #if(len(sys.argv) < 1):
  wb = load_workbook(basedir + '\\shulCloud\\yahrzeits.xlsx')
  #else:
  #    wb = load_workbook('./' + str(sys.argv[1]))
  return wb

def fixDays(sheet):
    for r in range(2, sheet.max_row):
        if(len(sheet.cell(row=r, column=4).value.strip()) > 0 and sheet.cell(row=r, column=4).value[1] == ' '):
            sheet.cell(row=r, column=4).value = '0' + sheet.cell(row=r, column=4).value

def swapRows(sheet, r1, r2):
    rHold = ""
    for c in range(1,9):
        rHold = sheet.cell(row=r1, column=c).value
        sheet.cell(row=r1,column=c).value = sheet.cell(row=r2,column=c).value
        sheet.cell(row=r2,column=c).value = rHold

def sortSheet(sheet):
    for r1 in range(1, sheet.max_row + 1):
        for r2 in range(r1+1, sheet.max_row + 1):
            if(sheet.cell(row=r1, column=4).value > sheet.cell(row=r2, column=4).value):
                swapRows(sheet, r1, r2)

def finishSheet(sheet, title):
    sheet.cell(row=1,column=2).value = title
    sheet.cell(row=1,column=5).value = (datetime.now()).strftime("%B %d, %Y")

    sheet.cell(row=1,column=2).font = Font(color=colors.BLACK, bold=True)
    sheet.cell(row=1,column=5).font = Font(color=colors.BLACK, bold=True)

    column = 2
    while column < 10:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 16
        column += 1

    sheet.column_dimensions[get_column_letter(3)].width = 12
    sheet.column_dimensions[get_column_letter(1)].width = 25
    sheet.column_dimensions[get_column_letter(2)].width = 25

def addBorderDivisions(sheet):
    underline = Border(left=Side(style='none'),
                     right=Side(style='none'),
                     top=Side(style='none'),
                     bottom=Side(style='double'))

    day = str(sheet.cell(row=2,column=4).value)[0:2]

    for r in range(3, sheet.max_row):
        if (day != str(sheet.cell(row=r,column=4).value)[0:2]):
            for s in range(1,9):
                sheet.cell(row=r-1, column=s).border = underline
                day = str(sheet.cell(row=r,column=4).value)[0:2]

def splitByGender(wbook):
    sheet = wbook['OneMonth']
    sheetM = wbook.create_sheet(title = 'Males')
    sheetM = wbook.active
    sheetF = wbook.create_sheet(title = 'Females')
    sheetF = wbook.active
    MaleRowCount = 1
    FemaleRowCount = 1

    for r in range(2, sheet.max_row):
        if(sheet.cell(row=r,column=6).value == 'Male'):
            sheetG = wbook['Males']
            MaleRowCount = MaleRowCount + 1
            count = MaleRowCount
        else:
            sheetG = wbook['Females']
            FemaleRowCount = FemaleRowCount + 1
            count = FemaleRowCount

        sheetG.cell(count, 1).value = sheet.cell(row=r,column=2).value
        sheetG.cell(count, 2).value = sheet.cell(row=r,column=5).value

        #if (str(sheet.cell(row=r,column=7).value) == "0000-00-00"):
        #    sheetG.cell(count, 3).value = ""
        #else:
        sheetG.cell(count, 3).value = sheet.cell(row=r,column=24).value
        sheetG.cell(count, 4).value = sheet.cell(row=r,column=8).value
        sheetG.cell(count, 5).value = sheet.cell(row=r,column=25).value
        sheetG.cell(count, 6).value = sheet.cell(row=r,column=28).value
        sheetG.cell(count, 7).value = sheet.cell(row=r,column=29).value
        sheetG.cell(count, 8).value = sheet.cell(row=r,column=33).value

def splitByGenderII(wbook):
    sheet = wbook['OneMonthII']
    sheetM = wbook.create_sheet(title = 'MalesII')
    sheetM = wbook.active
    sheetF = wbook.create_sheet(title = 'FemalesII')
    sheetF = wbook.active
    MaleRowCount = 1
    FemaleRowCount = 1

    for r in range(2, sheet.max_row):
        if(sheet.cell(row=r,column=6).value == 'Male'):
            sheetG = wbook['MalesII']
            MaleRowCount = MaleRowCount + 1
            count = MaleRowCount
        else:
            sheetG = wbook['FemalesII']
            FemaleRowCount = FemaleRowCount + 1
            count = FemaleRowCount

        sheetG.cell(count, 1).value = sheet.cell(row=r,column=2).value
        sheetG.cell(count, 2).value = sheet.cell(row=r,column=5).value

        #if (str(sheet.cell(row=r,column=7).value) == "0000-00-00"):
        #    sheetG.cell(count, 3).value = ""
        #else:
        sheetG.cell(count, 3).value = sheet.cell(row=r,column=24).value
        sheetG.cell(count, 4).value = sheet.cell(row=r,column=8).value
        sheetG.cell(count, 5).value = sheet.cell(row=r,column=25).value
        sheetG.cell(count, 6).value = sheet.cell(row=r,column=28).value
        sheetG.cell(count, 7).value = sheet.cell(row=r,column=29).value
        sheetG.cell(count, 8).value = sheet.cell(row=r,column=33).value

def stripMonth(wbook):
    global adarII
    newRow = 1
    newRowII = 1
    sheet = wbook[wbook.sheetnames[0]]
    sheetNew = wbook.create_sheet(title = 'OneMonth')
    #sheetNew = wbook.active
    mName = getNextMonthName().lower()
    if mName == "iyar":
        mName = "iyyar"
        
    if mName.find("adar") > -1:
        mName = "adar"
        sheetNewII = wbook.create_sheet(title = 'OneMonthII')
        adarII = True
    else:
        if mName == "shvat":
            mName = "shevat"
        else:
            if mName[-1] == 't' or mName[-1] == 's':
                mName = mName[0:-1]

    for r in range(2, sheet.max_row):
        if mName.find("adar") > -1:
            if sheet.cell(row=r,column=8).value.lower().find("ii") > -1:
                for s in range(1, 40):
                    sheetNewII.cell(row=newRowII, column=s).value = sheet.cell(row=r,column=s).value
                newRowII += 1
                continue

        if(sheet.cell(row=r,column=8).value.lower().find(mName) > -1):
            for s in range(1, 40):
                sheetNew.cell(row=newRow, column=s).value = sheet.cell(row=r,column=s).value
            newRow += 1

wbook = openwb()
stripMonth(wbook)
splitByGender(wbook)
sheet = wbook['Males']
fixDays(sheet)
sortSheet(sheet)
finishSheet(sheet, "Males")
addBorderDivisions(sheet)
if adarII:
    splitByGenderII(wbook)
    sheet = wbook['MalesII']
    fixDays(sheet)
    sortSheet(sheet)
    finishSheet(sheet, "MalesII")
    addBorderDivisions(sheet)

sheet = wbook['Females']
fixDays(sheet)
sortSheet(sheet)
finishSheet(sheet, "Females")
addBorderDivisions(sheet)

if adarII:
    sheet = wbook['FemalesII']
    fixDays(sheet)
    sortSheet(sheet)
    finishSheet(sheet, "FemalesII")
    addBorderDivisions(sheet)

wbook.save(basedir + '\\shulCloud\\new.xlsx')
